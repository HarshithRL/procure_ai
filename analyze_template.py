#!/usr/bin/env python3
"""Analyze the comparison template structure."""

import openpyxl
from pathlib import Path

template_path = Path("C:/Users/HarshithR/Downloads/Comparison Template (1).xlsx")

wb = openpyxl.load_workbook(template_path)

print("Sheet names:", wb.sheetnames)
print("\n" + "="*80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n### SHEET: {sheet_name} ###")
    print(f"Dimensions: {ws.dimensions}")
    print(f"\nFirst 25 rows (sample):")
    
    for i, row in enumerate(ws.iter_rows(max_row=25, values_only=True), 1):
        print(f"Row {i}: {row}")
