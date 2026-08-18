"""Build vendor comparison matrices in Excel format.

Incorporates the official Comparison Template structure with procurement metrics.
Supports multiple sheets: BAFO (Best and Final Offer), Scorecard, Cost Analysis, Risk Assessment.
Includes vendor logos and branding.
"""

from __future__ import annotations

import io
import json
import logging
import sqlite3
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import requests

logger = logging.getLogger(__name__)


class ComparisonMatrixBuilder:
    """Build vendor comparison Excel workbooks with full procurement metrics."""
    
    def __init__(self, project_id: int, db_path: str):
        self.project_id = project_id
        self.db_path = db_path
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Styling
        self.header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.subheader_fill = PatternFill(start_color="E7F0FF", end_color="E7F0FF", fill_type="solid")
        self.subheader_font = Font(bold=True, size=11)
        self.metric_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        self.metric_font = Font(size=10)
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.logo_cache = {}  # Cache for generated/fetched logos
    
    def _generate_logo(self, vendor_name: str) -> bytes:
        """Generate a simple logo placeholder PNG for a vendor (cached)."""
        if vendor_name in self.logo_cache:
            return self.logo_cache[vendor_name]
        
        try:
            # Create a colorful logo with vendor initials
            img = Image.new('RGB', (200, 200), color=(70, 130, 180))  # Steel blue background
            draw = ImageDraw.Draw(img)
            
            # Add vendor name
            # Try to use a default font; fall back to default if not available
            try:
                font = ImageFont.truetype("arial.ttf", 36)
            except:
                font = ImageFont.load_default()
            
            # Get initials
            initials = ''.join([word[0].upper() for word in vendor_name.split() if word])[:2]
            
            # Add text
            text_bbox = draw.textbbox((0, 0), initials, font=font)
            text_width = text_bbox[2] - text_bbox[0]
            text_height = text_bbox[3] - text_bbox[1]
            x = (200 - text_width) // 2
            y = (200 - text_height) // 2
            
            draw.text((x, y), initials, font=font, fill=(255, 255, 255))
            
            # Add border
            draw.rectangle([(5, 5), (195, 195)], outline=(255, 255, 255), width=3)
            
            # Save to bytes
            logo_bytes = io.BytesIO()
            img.save(logo_bytes, format='PNG')
            logo_bytes.seek(0)
            result = logo_bytes.getvalue()
            
            self.logo_cache[vendor_name] = result
            return result
        except Exception as e:
            logger.warning(f"Failed to generate logo for {vendor_name}: {e}")
            return None
    
    def _get_vendors_from_db(self) -> list[dict]:
        """Fetch all vendors for this project from SQLite."""
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, node_type, label, slug, metadata
                FROM nodes
                WHERE project_id = ? AND node_type = 'vendor'
                ORDER BY label
            """, (self.project_id,))
            
            vendors = []
            for row in cursor.fetchall():
                metadata = json.loads(row['metadata'] or '{}')
                vendors.append({
                    'id': row['id'],
                    'name': row['label'],
                    'slug': row['slug'],
                    'domain': metadata.get('domain'),
                    'location': metadata.get('location'),
                    'certifications': metadata.get('certifications', []),
                    'quality_summary': metadata.get('quality_summary'),
                    'logo_url': metadata.get('logo_url'),  # Optional: if vendor has a logo URL
                })
            
            conn.close()
            return vendors
        except Exception as e:
            logger.error(f"Failed to fetch vendors from DB: {e}")
            return []
    
    def _get_offerings_from_db(self, vendor_id: Optional[str] = None) -> list[dict]:
        """Fetch offerings for this project, optionally filtered by vendor."""
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            if vendor_id:
                cursor.execute("""
                    SELECT n.id, n.label, n.slug, n.metadata,
                           (SELECT MAX(confidence) FROM edges e WHERE e.source_id = ? AND e.target_id = n.id) as link_confidence
                    FROM nodes n
                    WHERE n.project_id = ? AND n.node_type = 'offering'
                    AND EXISTS (SELECT 1 FROM edges e WHERE e.source_id = ? AND e.target_id = n.id)
                    ORDER BY n.label
                """, (vendor_id, self.project_id, vendor_id))
            else:
                cursor.execute("""
                    SELECT id, label, slug, metadata
                    FROM nodes
                    WHERE project_id = ? AND node_type = 'offering'
                    ORDER BY label
                """, (self.project_id,))
            
            offerings = []
            for row in cursor.fetchall():
                metadata = json.loads(row['metadata'] or '{}')
                offerings.append({
                    'id': row['id'],
                    'name': row['label'],
                    'slug': row['slug'],
                    'vendor_name': metadata.get('vendor_name'),
                    'delivery_timeline': metadata.get('delivery_timeline'),
                    'support_model': metadata.get('support_model'),
                })
            
            conn.close()
            return offerings
        except Exception as e:
            logger.error(f"Failed to fetch offerings from DB: {e}")
            return []
    
    def _get_offer_lines_from_db(self, offering_id: Optional[str] = None) -> list[dict]:
        """Fetch offer lines (pricing) for this project."""
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            if offering_id:
                cursor.execute("""
                    SELECT n.id, n.label, n.slug, n.metadata
                    FROM nodes n
                    WHERE n.project_id = ? AND n.node_type = 'offer_line'
                    AND EXISTS (SELECT 1 FROM edges e WHERE e.source_id = ? AND e.target_id = n.id)
                    ORDER BY n.label
                """, (self.project_id, offering_id))
            else:
                cursor.execute("""
                    SELECT id, label, slug, metadata
                    FROM nodes
                    WHERE project_id = ? AND node_type = 'offer_line'
                    ORDER BY label
                """, (self.project_id,))
            
            lines = []
            for row in cursor.fetchall():
                metadata = json.loads(row['metadata'] or '{}')
                lines.append({
                    'id': row['id'],
                    'name': row['label'],
                    'slug': row['slug'],
                    'vendor_name': metadata.get('vendor_name'),
                    'article': metadata.get('article'),
                    'quantity': metadata.get('quantity'),
                    'unit_price': metadata.get('unit_price'),
                    'currency': metadata.get('currency', 'EUR'),
                    'total_cost': metadata.get('total_cost'),
                    'payment_terms': metadata.get('payment_terms'),
                })
            
            conn.close()
            return lines
        except Exception as e:
            logger.error(f"Failed to fetch offer lines from DB: {e}")
            return []
    
    def build_bafo_sheet(self):
        """Build the BAFO (Best and Final Offer) comparison sheet based on template."""
        ws = self.wb.create_sheet("BAFO Comparison")
        
        vendors = self._get_vendors_from_db()[:4]  # Max 4 vendors in template
        if not vendors:
            logger.warning(f"No vendors found for project {self.project_id}")
            vendors = [{"name": f"Vendor {i+1}", "slug": f"vendor-{i+1}", "logo_url": None} for i in range(2)]
        
        # Row 1: Logo row
        ws.row_dimensions[1].height = 80  # Tall enough for logos
        ws['A1'] = "Metric Category"
        ws['B1'] = "Type"
        
        for idx, vendor in enumerate(vendors, start=3):
            col_letter = get_column_letter(idx)
            
            # Try to add logo
            try:
                logo_bytes = None
                
                # Try fetching from URL if available
                if vendor.get('logo_url'):
                    try:
                        response = requests.get(vendor['logo_url'], timeout=5)
                        if response.status_code == 200:
                            logo_bytes = response.content
                    except Exception as e:
                        logger.debug(f"Failed to fetch logo from URL: {e}")
                
                # Generate placeholder logo if no URL or fetch failed
                if not logo_bytes:
                    logo_bytes = self._generate_logo(vendor['name'])
                
                # Insert logo as image
                if logo_bytes:
                    logo_io = io.BytesIO(logo_bytes)
                    img = XLImage(logo_io)
                    img.width = 120
                    img.height = 80
                    ws.add_image(img, f'{col_letter}1')
            except Exception as e:
                logger.warning(f"Failed to add logo for {vendor['name']}: {e}")
        
        # Row 2: Vendor name headers
        ws.row_dimensions[2].height = 25
        ws['A2'] = "Metric Category"
        ws['B2'] = "Type"
        
        for idx, vendor in enumerate(vendors, start=3):
            ws.cell(row=2, column=idx, value=vendor['name'])
        
        # Style header rows
        for col in range(1, len(vendors) + 3):
            for row_num in [1, 2]:
                cell = ws.cell(row=row_num, column=col)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = self.border
        
        # Cost section (Year 1, 2, 3)
        row = 4  # Start after logo header rows
        cost_categories = [
            ("Year 1 Costs", [
                ("Solution Costs (Software, Materials)", "Recurring", "EUR"),
                ("Service/Implementation Costs", "One-off", "EUR"),
                ("Etex Mandays (standard rate 800 EUR/day)", "Recurring", "# of days × 800"),
                ("Total TCO Year 1", None, "SUM"),
            ]),
            ("Year 2 Costs", [
                ("Solution Costs (Software, Materials)", "Recurring", "EUR"),
                ("Service/Implementation Costs (if applicable)", "One-off", "EUR"),
                ("Etex Mandays (if applicable)", "Recurring", "# of days × 800"),
                ("Total TCO Year 2", None, "SUM (indexed)"),
            ]),
            ("Year 3 Costs", [
                ("Solution Costs (Software, Materials)", "Recurring", "EUR"),
                ("Service/Implementation Costs (if applicable)", "One-off", "EUR"),
                ("Etex Mandays (if applicable)", "Recurring", "# of days × 800"),
                ("Total TCO Year 3", None, "SUM (indexed)"),
            ]),
        ]
        
        for category_name, items in cost_categories:
            # Category header
            ws.cell(row=row, column=1, value=category_name)
            ws.cell(row=row, column=1).fill = self.subheader_fill
            ws.cell(row=row, column=1).font = self.subheader_font
            row += 1
            
            for metric, metric_type, default_value in items:
                ws.cell(row=row, column=1, value=metric)
                ws.cell(row=row, column=2, value=metric_type)
                
                for vendor_idx in range(len(vendors)):
                    col = vendor_idx + 3
                    ws.cell(row=row, column=col, value=default_value)
                    ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
                
                row += 1
        
        # Total TCO across all years
        ws.cell(row=row, column=1, value="Total TCO (3/5 Years)")
        ws.cell(row=row, column=1).fill = self.subheader_fill
        ws.cell(row=row, column=1).font = self.subheader_font
        for vendor_idx in range(len(vendors)):
            col = vendor_idx + 3
            ws.cell(row=row, column=col, value="SUM of annual costs")
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="right")
        row += 2
        
        # Scoring & Assessment section
        assessment_metrics = [
            ("Score Card Score", "0-100%", "Weighted evaluation"),
            ("Indexation (yearly %)", "Percentage", "Escalation rate"),
            ("Indicative Blended Rate", "EUR/day", "Service cost per day"),
            ("Vendor Mandays (Implementation)", "Days", "Vendor effort"),
            ("Etex Mandays (Internal Effort)", "Days", "Etex effort"),
            ("Planning Timeline", "Weeks", "Go-live weeks"),
            ("EcoVadis Score", "0-100", "Sustainability rating"),
            ("Credit Score", "0-100", "Financial health"),
            ("Architecture Assessment", "0-100", "Technical fit"),
            ("Security Assessment", "0-100", "Security posture"),
            ("Delivery Risk", "Low/Med/High", "Risk level"),
            ("Support Quality", "1-5", "Support rating"),
            ("Certification Status", "Yes/No", "ISO/Compliance"),
            ("Reference Availability", "Yes/No", "Customer refs"),
        ]
        
        ws.cell(row=row, column=1, value="Scorecards & Assessment")
        ws.cell(row=row, column=1).fill = self.subheader_fill
        ws.cell(row=row, column=1).font = self.subheader_font
        row += 1
        
        for metric_name, metric_type, description in assessment_metrics:
            ws.cell(row=row, column=1, value=metric_name)
            ws.cell(row=row, column=2, value=metric_type)
            
            for vendor_idx in range(len(vendors)):
                col = vendor_idx + 3
                ws.cell(row=row, column=col, value="")  # Empty for user to fill
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        for idx in range(len(vendors)):
            col_letter = get_column_letter(idx + 3)
            ws.column_dimensions[col_letter].width = 20
        
        # Set row height for header
        ws.row_dimensions[1].height = 30
    
    def build_scorecard_sheet(self):
        """Build detailed scorecard with weighted scoring."""
        ws = self.wb.create_sheet("Scorecard")
        
        vendors = self._get_vendors_from_db()[:4]
        if not vendors:
            return
        
        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Vendor Comparison Scorecard - Project {self.project_id}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].fill = self.header_fill
        ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        row = 4
        # Header
        ws.cell(row=row, column=1, value="Evaluation Criterion")
        ws.cell(row=row, column=2, value="Weight %")
        for idx, vendor in enumerate(vendors, start=3):
            ws.cell(row=row, column=idx, value=f"{vendor['name']} Score")
        
        # Scoring criteria (extracted from procurement domain knowledge)
        scoring_criteria = [
            ("Technical Fit", 25),
            ("Cost (TCO)", 30),
            ("Implementation Timeline", 15),
            ("Vendor Stability & Viability", 10),
            ("Support & Service Quality", 10),
            ("Security & Compliance", 10),
        ]
        
        row += 1
        total_weight = 0
        weighted_scores = {vendor['name']: 0 for vendor in vendors}
        
        for criterion, weight in scoring_criteria:
            ws.cell(row=row, column=1, value=criterion)
            ws.cell(row=row, column=2, value=f"{weight}%")
            total_weight += weight
            
            for vendor_idx, vendor in enumerate(vendors, start=3):
                ws.cell(row=row, column=vendor_idx, value="")  # User fills in 0-100
            
            row += 1
        
        # Total row
        ws.cell(row=row, column=1, value="TOTAL WEIGHTED SCORE")
        ws.cell(row=row, column=1).fill = self.header_fill
        ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")
        
        ws.cell(row=row, column=2, value=f"{total_weight}%")
        ws.cell(row=row, column=2).fill = self.header_fill
        ws.cell(row=row, column=2).font = Font(bold=True, color="FFFFFF")
        
        for vendor_idx in range(len(vendors)):
            col = vendor_idx + 3
            ws.cell(row=row, column=col, value="")  # Calculated
            ws.cell(row=row, column=col).fill = self.header_fill
            ws.cell(row=row, column=col).font = Font(bold=True, color="FFFFFF")
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 12
        for idx in range(len(vendors)):
            col_letter = get_column_letter(idx + 3)
            ws.column_dimensions[col_letter].width = 20
    
    def build_requirements_matrix_sheet(self):
        """Build requirements vs vendors matrix."""
        ws = self.wb.create_sheet("Requirements Matrix")
        
        vendors = self._get_vendors_from_db()[:4]
        if not vendors:
            return
        
        # Fetch requirements from DB
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, label, slug, metadata
                FROM nodes
                WHERE project_id = ? AND node_type = 'requirement'
                ORDER BY label
            """, (self.project_id,))
            
            requirements = []
            for row in cursor.fetchall():
                metadata = json.loads(row['metadata'] or '{}')
                requirements.append({
                    'name': row['label'],
                    'category': metadata.get('category', 'functional'),
                    'measurable': metadata.get('measurable', False),
                    'metric': metadata.get('metric'),
                })
            
            conn.close()
        except Exception as e:
            logger.error(f"Failed to fetch requirements: {e}")
            requirements = []
        
        if not requirements:
            requirements = [
                {"name": "Response Time", "category": "Technical", "measurable": True, "metric": "<100ms"},
                {"name": "Availability", "category": "Technical", "measurable": True, "metric": ">99.5%"},
                {"name": "Support Hours", "category": "Support", "measurable": True, "metric": "24/7"},
            ]
        
        # Header row
        row = 1
        ws.cell(row=row, column=1, value="Requirement")
        ws.cell(row=row, column=2, value="Category")
        ws.cell(row=row, column=3, value="Measurable")
        ws.cell(row=row, column=4, value="Target Metric")
        
        for idx, vendor in enumerate(vendors, start=5):
            ws.cell(row=row, column=idx, value=f"{vendor['name']}\nCompliance")
        
        # Style header
        for col in range(1, len(vendors) + 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        ws.row_dimensions[row].height = 30
        
        # Requirements rows
        row += 1
        for req in requirements:
            ws.cell(row=row, column=1, value=req['name'])
            ws.cell(row=row, column=2, value=req['category'])
            ws.cell(row=row, column=3, value="Yes" if req['measurable'] else "No")
            ws.cell(row=row, column=4, value=req['metric'] or "")
            
            for vendor_idx in range(len(vendors)):
                col = vendor_idx + 5
                ws.cell(row=row, column=col, value="✓")  # User marks compliance
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        for idx in range(len(vendors)):
            col_letter = get_column_letter(idx + 5)
            ws.column_dimensions[col_letter].width = 18
    
    def build_risk_assessment_sheet(self):
        """Build risk assessment matrix."""
        ws = self.wb.create_sheet("Risk Assessment")
        
        vendors = self._get_vendors_from_db()[:4]
        if not vendors:
            return
        
        # Fetch risks from DB
        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            
            cursor.execute("""
                SELECT id, label, slug, metadata
                FROM nodes
                WHERE project_id = ? AND node_type = 'risk'
                ORDER BY label
            """, (self.project_id,))
            
            risks = []
            for row in cursor.fetchall():
                metadata = json.loads(row['metadata'] or '{}')
                risks.append({
                    'name': row['label'],
                    'severity': metadata.get('severity', 'medium'),
                    'related_vendors': metadata.get('related_vendors', []),
                })
            
            conn.close()
        except Exception as e:
            logger.error(f"Failed to fetch risks: {e}")
            risks = []
        
        if not risks:
            risks = [
                {"name": "Implementation Delay Risk", "severity": "high"},
                {"name": "Vendor Financial Viability", "severity": "medium"},
                {"name": "Technical Compatibility", "severity": "medium"},
            ]
        
        # Header
        row = 1
        ws.cell(row=row, column=1, value="Risk Item")
        ws.cell(row=row, column=2, value="Severity")
        for idx, vendor in enumerate(vendors, start=3):
            ws.cell(row=row, column=idx, value=f"{vendor['name']}\nRisk Level")
        
        # Style header
        for col in range(1, len(vendors) + 3):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        ws.row_dimensions[row].height = 25
        
        # Risk rows
        row += 1
        for risk in risks:
            ws.cell(row=row, column=1, value=risk['name'])
            ws.cell(row=row, column=2, value=risk['severity'].upper())
            
            # Color severity
            severity_color = {"critical": "FF0000", "high": "FF6600", "medium": "FFCC00", "low": "00CC00"}
            ws.cell(row=row, column=2).fill = PatternFill(
                start_color=severity_color.get(risk['severity'].lower(), "FFFFFF"),
                end_color=severity_color.get(risk['severity'].lower(), "FFFFFF"),
                fill_type="solid"
            )
            
            for vendor_idx in range(len(vendors)):
                col = vendor_idx + 3
                ws.cell(row=row, column=col, value="")  # Low/Medium/High
            
            row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 15
        for idx in range(len(vendors)):
            col_letter = get_column_letter(idx + 3)
            ws.column_dimensions[col_letter].width = 18
    
    def build_summary_sheet(self):
        """Build a summary overview sheet with vendor logos, key metrics, and recommendation."""
        ws = self.wb.create_sheet("Summary", 0)  # Insert as first sheet
        
        vendors = self._get_vendors_from_db()[:4]
        if not vendors:
            return
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f"VENDOR COMPARISON SUMMARY"
        ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
        ws['A1'].fill = self.header_fill
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        
        # Project info
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Project: {self.project_id} | Date: {datetime.now().strftime('%Y-%m-%d')}"
        ws['A2'].font = Font(size=10, italic=True)
        ws['A2'].fill = PatternFill(start_color="E7F0FF", end_color="E7F0FF", fill_type="solid")
        
        # Vendor logos section
        ws.merge_cells('A3:F3')
        ws['A3'] = "Vendors Under Consideration"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].fill = self.subheader_fill
        
        # Add logos
        ws.row_dimensions[4].height = 100
        col_positions = [('A', 0), ('C', 1), ('E', 2)]  # Wider spacing
        
        for col_letter, vendor_idx in col_positions:
            if vendor_idx < len(vendors):
                vendor = vendors[vendor_idx]
                
                try:
                    logo_bytes = None
                    
                    # Try fetching from URL
                    if vendor.get('logo_url'):
                        try:
                            response = requests.get(vendor['logo_url'], timeout=5)
                            if response.status_code == 200:
                                logo_bytes = response.content
                        except Exception as e:
                            logger.debug(f"Failed to fetch logo: {e}")
                    
                    # Generate placeholder
                    if not logo_bytes:
                        logo_bytes = self._generate_logo(vendor['name'])
                    
                    if logo_bytes:
                        logo_io = io.BytesIO(logo_bytes)
                        img = XLImage(logo_io)
                        img.width = 150
                        img.height = 100
                        ws.add_image(img, f'{col_letter}4')
                except Exception as e:
                    logger.warning(f"Failed to add logo: {e}")
        
        # Vendor names below logos
        row = 6
        ws.row_dimensions[row].height = 20
        
        for col_idx, (col_letter, vendor_idx) in enumerate(col_positions):
            if vendor_idx < len(vendors):
                vendor = vendors[vendor_idx]
                cell = ws[f'{col_letter}{row}']
                cell.value = vendor['name']
                cell.font = Font(bold=True, size=11)
                cell.alignment = Alignment(horizontal="center")
                cell.fill = self.subheader_fill
        
        # Key metrics table
        row = 8
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Key Comparison Metrics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.subheader_fill
        
        row += 1
        
        # Metrics header
        ws[f'A{row}'] = "Metric"
        ws[f'B{row}'] = f"{vendors[0]['name']}" if len(vendors) > 0 else "Vendor 1"
        ws[f'C{row}'] = f"{vendors[1]['name']}" if len(vendors) > 1 else "Vendor 2"
        ws[f'D{row}'] = f"{vendors[2]['name']}" if len(vendors) > 2 else "Vendor 3"
        ws[f'E{row}'] = f"{vendors[3]['name']}" if len(vendors) > 3 else "Vendor 4"
        ws[f'F{row}'] = "Winner"
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col}{row}']
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = self.border
        
        # Sample metrics to populate
        metrics = [
            ("Total TCO (3 years)", "", "", "", "", ""),
            ("Implementation Timeline", "", "", "", "", ""),
            ("Score Card Rating", "", "", "", "", ""),
            ("Vendor Viability", "", "", "", "", ""),
            ("Support Quality", "", "", "", "", ""),
            ("Security/Compliance", "", "", "", "", ""),
        ]
        
        row += 1
        for metric_name, *values in metrics:
            ws[f'A{row}'] = metric_name
            for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E']):
                ws[f'{col_letter}{row}'] = ""
            ws[f'F{row}'] = ""
            
            ws[f'A{row}'].fill = self.metric_fill
            ws[f'A{row}'].font = Font(bold=True)
            
            for col_letter in ['B', 'C', 'D', 'E', 'F']:
                ws[f'{col_letter}{row}'].border = self.border
            
            row += 1
        
        # Recommendation section
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws[f'A{row}'] = "Recommendation"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = self.subheader_fill
        
        row += 1
        ws.merge_cells(f'A{row}:F{row+2}')
        ws[f'A{row}'] = "Recommended Vendor: [To be filled in]\n\nRationale: [Document key decision factors]"
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[row].height = 60
        
        # Set column widths
        ws.column_dimensions['A'].width = 25
        for col_letter in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 18
    
    def build(self) -> bytes:
        """Build and return Excel workbook as bytes."""
        try:
            self.build_summary_sheet()  # First sheet with logos
            self.build_bafo_sheet()
            self.build_scorecard_sheet()
            self.build_requirements_matrix_sheet()
            self.build_risk_assessment_sheet()
            
            # Write to bytes buffer
            buffer = io.BytesIO()
            self.wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
        except Exception as e:
            logger.error(f"Failed to build workbook: {e}")
            raise
