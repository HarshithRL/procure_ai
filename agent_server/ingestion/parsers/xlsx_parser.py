"""XLSX parsing via openpyxl: per-sheet cell ranges (pricing matrices)."""

from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from agent_server.knowledge_graph.schema import Chunk

_MAX_HEADER_SCAN_ROWS = 5


def parse_xlsx(file_content: bytes, document_id: str) -> list[Chunk]:
    """Parse every sheet in an XLSX workbook into one `Chunk` per data row."""
    workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True, read_only=True)
    try:
        chunks: list[Chunk] = []
        for sheet in workbook.worksheets:
            chunks.extend(_parse_sheet(sheet, document_id))
        return chunks
    finally:
        workbook.close()


def _parse_sheet(sheet: Worksheet, document_id: str) -> list[Chunk]:
    chunks: list[Chunk] = []
    rows = list(sheet.iter_rows(values_only=False))
    if not rows:
        return chunks

    header_row_index, headers = _find_header_row(rows)
    data_rows = rows[header_row_index + 1 :] if header_row_index is not None else rows
    first_data_row_number = (header_row_index + 2) if header_row_index is not None else 1

    for offset, row in enumerate(data_rows):
        row_number = first_data_row_number + offset
        values = [cell.value for cell in row]
        if all(value is None for value in values):
            continue
        last_col = get_column_letter(len(values)) if values else "A"
        locator = f"{sheet.title}!A{row_number}:{last_col}{row_number}"
        chunks.append(
            Chunk(
                document_id=document_id,
                locator=locator,
                text=_row_text(headers, values),
                metadata={"sheet": sheet.title, "row": row_number, "headers": headers},
            )
        )
    return chunks


def _find_header_row(rows: list) -> tuple[int | None, list[str]]:
    for row_index in range(min(_MAX_HEADER_SCAN_ROWS, len(rows))):
        values = [cell.value for cell in rows[row_index]]
        non_empty = [value for value in values if value not in (None, "")]
        if len(non_empty) >= max(2, len(values) // 2):
            headers = [str(value) if value is not None else f"col_{i + 1}" for i, value in enumerate(values)]
            return row_index, headers
    return None, []


def _row_text(headers: list[str], values: list) -> str:
    if headers:
        pairs = [f"{header}={value}" for header, value in zip(headers, values) if value is not None]
    else:
        pairs = [str(value) for value in values if value is not None]
    return " | ".join(pairs)
